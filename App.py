"""
ASCENT GHG Viewer — Flask Application
Parses an uploaded ASCENT Excel (.xlsm/.xlsx) file and renders:
  1. Upload page
  2. BAU Emission Profile
  3. Target Setting
  4. Dashboard-BAU
  5. Emission Reduction Graph
"""

import os, io, json, traceback
from flask import (Flask, render_template, request, redirect,
                   url_for, session, jsonify, flash)
from openpyxl import load_workbook
import pandas as pd

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "ascent-ghg-secret-2025")
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024   # 50 MB limit

UPLOAD_FOLDER = "/tmp/ascent_uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)


# ═══════════════════════════════════════════════════════
#  EXCEL PARSING HELPERS
# ═══════════════════════════════════════════════════════

def safe_float(v, default=0.0):
    try:
        if v is None or v == "" or str(v).strip() in ("#REF!", "#VALUE!", "#NAME?", "Select Activity"):
            return default
        return float(v)
    except (TypeError, ValueError):
        return default

def safe_int(v, default=0):
    try:
        return int(safe_float(v, default))
    except:
        return default

def ws_rows(wb, sheet_name):
    """Return all non-empty rows from a worksheet as list of tuples."""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    result = []
    for row in ws.iter_rows(values_only=True):
        if any(c is not None for c in row):
            result.append(row)
    return result


def parse_basic_info(wb):
    """Extract from 'A. Basic Information' sheet."""
    rows = ws_rows(wb, "A. Basic Information")
    info = {}
    year_table = []

    in_year_table = False
    for row in rows:
        c = [r for r in row if r is not None]
        if not c:
            continue
        # Detect year table header
        if any("Base Year and Target Years" in str(v) for v in c):
            in_year_table = True
            continue
        if in_year_table and any(str(v) in ("Base Year","Interim Year 1","Interim Year 2","Target Year") for v in row):
            label = next((str(v) for v in row if str(v) in ("Base Year","Interim Year 1","Interim Year 2","Target Year")), None)
            vals = [safe_float(v) for v in row if isinstance(v, (int, float))]
            if label and len(vals) >= 4:
                year_table.append({
                    "label": label,
                    "year": safe_int(vals[0]),
                    "pop_growth": vals[1],
                    "gdp_growth": vals[2],
                    "growth_factor": vals[3],
                })

        # Parse basic key-value pairs
        for i, v in enumerate(row):
            if v is None:
                continue
            sv = str(v).strip()
            # Look for label → value in adjacent columns
            nxt = row[i+1] if i+1 < len(row) else None
            if sv == "State":              info["state"]        = str(nxt or "")
            elif sv == "Government Tier": info["tier"]         = str(nxt or "")
            elif sv == "Area Limit":      info["area_sqkm"]    = safe_float(nxt)
            elif sv == "Population":      info["population"]   = safe_float(nxt)
            elif sv == "Population Growth Rate": info["pop_growth_rate"] = safe_float(nxt)
            elif sv == "Avg. Annual Rainfall":   info["rainfall"]       = safe_float(nxt)
            elif sv == "Avg. Minimum Temprature":info["temp_min"]       = safe_float(nxt)
            elif sv == "Avg. Maximum Temprature":info["temp_max"]       = safe_float(nxt)
            elif sv == "City Average Temperature":info["temp_avg"]      = safe_float(nxt)
            elif sv == "City Climate Zone (as per NBC 2016)": info["climate_zone"] = str(nxt or "")
            elif sv == "City GDP":        info["gdp"]          = safe_float(nxt)
            elif sv == "GDP Growth rate": info["gdp_growth"]   = safe_float(nxt)
            elif sv == "Name" and i > 0:  info["city_name"]    = str(nxt or "")

    info["year_table"] = year_table
    # Extract base/target years from year_table
    for yt in year_table:
        if yt["label"] == "Base Year":    info["base_year"]    = yt["year"]
        if yt["label"] == "Target Year":  info["target_year"]  = yt["year"]
        if yt["label"] == "Interim Year 1": info["interim1"]   = yt["year"]
        if yt["label"] == "Interim Year 2": info["interim2"]   = yt["year"]
    return info


def parse_bau_scenario(wb):
    """
    Extract BAU data from 'BAU Scenario' sheet.
    Returns list of subsector rows with base/interim1/interim2/target emissions.
    Also returns per-capita and growth metadata.
    """
    rows = ws_rows(wb, "BAU Scenario")
    sector_data = []
    meta = {"years": [], "growth_rates": {}}
    years_found = []

    for row in rows:
        # Detect year columns row: contains 2025/2030/2040/2050
        year_vals = [v for v in row if isinstance(v, (int, float)) and 2020 <= v <= 2060]
        if len(year_vals) >= 3 and not years_found:
            years_found = [int(y) for y in year_vals[:4]]
            meta["years"] = years_found
            continue

        # Detect Population/GDP Growth Rate rows
        for v in row:
            if str(v or "").strip() == "Population Growth Rate (%)":
                rates = [safe_float(x) for x in row if isinstance(x, (int, float)) and 0 < x < 1]
                meta["growth_rates"]["population"] = rates
            if str(v or "").strip() == "GDP (₹ Crore)":
                vals = [safe_float(x) for x in row if isinstance(x, (int, float)) and x > 100]
                meta["gdp_series"] = vals

        # Parse sector/subsector rows
        sector = None
        subsector = None
        for i, v in enumerate(row):
            sv = str(v or "").strip()
            if sv in ("Energy Sector","Transport","Waste","AFOLU","IPPU"):
                sector = sv
                break
        if sector:
            nums = [safe_float(v) for v in row if isinstance(v, (int, float)) and v != 0]
            if len(nums) >= 3:
                # Row: sector, subsector, base, per_capita, 2030, 2040, 2050
                sector_data.append({"sector": sector, "subsector": "Total", "values": nums[:4]})

        # Check for subsector rows (None, None, subsector_name, ...)
        if row[2] is None and row[3] is not None:
            subsector = str(row[3]).strip()
            if subsector and subsector not in ("", "Total Emissions", "Not Considered"):
                nums = []
                for v in row[4:]:
                    if isinstance(v, (int, float)) and abs(v) > 0.001:
                        nums.append(safe_float(v))
                if len(nums) >= 3:
                    sector_data.append({"sector": "?", "subsector": subsector, "values": nums[:4]})

    return sector_data, meta


def parse_dashboard_bau(wb):
    """
    Parse 'Dashboard- BAU-City' sheet — the most reliable source for BAU data.
    Returns structured list: [{sector, subsector, base, y2030, y2040, y2050}, ...]
    """
    rows = ws_rows(wb, "Dashboard- BAU-City")
    data = []
    current_sector = None

    for row in rows:
        # Identify sector label from col D (index 3)
        sector_label = str(row[3] or "").strip() if len(row) > 3 else ""
        subsector_label = str(row[4] or "").strip() if len(row) > 4 else ""

        if sector_label in ("Energy Sector", "Transport", "Waste", "AFOLU", "IPPU"):
            current_sector = sector_label

        if not subsector_label or subsector_label in ("Subsector", "", "Total Emissions (MtCO2e)"):
            continue

        # Extract numeric values — cols 6,8,10,12 (base, 2030, 2040, 2050)
        def gc(idx):
            return safe_float(row[idx]) if idx < len(row) else 0.0

        base  = gc(6)
        y2030 = gc(8)
        y2040 = gc(10)
        y2050 = gc(12)

        if base == 0 and y2030 == 0 and y2040 == 0 and y2050 == 0:
            continue

        data.append({
            "sector":    current_sector or sector_label,
            "subsector": subsector_label,
            "base":      base,
            "y2030":     y2030,
            "y2040":     y2040,
            "y2050":     y2050,
        })

    return data


def parse_target_setting(wb):
    """
    Parse 'Target Setting ' sheet.
    Returns: {year: {bau, target_pct, target_abs, allowable}, ...}
    """
    rows = ws_rows(wb, "Target Setting ")
    targets = {}

    for row in rows:
        year_val = row[2] if len(row) > 2 else None
        if not isinstance(year_val, (int, float)):
            continue
        year = int(year_val)
        if year not in (2025, 2030, 2040, 2050):
            continue

        bau         = safe_float(row[3] if len(row) > 3 else 0)
        tgt_abs_raw = safe_float(row[4] if len(row) > 4 else 0)
        tgt_pct     = safe_float(row[5] if len(row) > 5 else 0)
        allowable   = safe_float(row[6] if len(row) > 6 else 0)

        targets[year] = {
            "bau":         bau,
            "target_abs":  tgt_abs_raw,
            "target_pct":  tgt_pct,
            "allowable":   allowable,
        }

    return targets


def parse_emission_reduction(wb):
    """
    Parse 'Emission Reduction- Graph' sheet.
    Returns: list of subsector rows with BAU, E&P, Ambitious for each milestone year.
    Also returns summary timeline rows (BAU, E&P, High-Ambition, Target).
    """
    rows = ws_rows(wb, "Emission Reduction- Graph")
    subsectors = []
    timeline = {}
    header_found = False
    summary_start = False

    for row in rows:
        # Detect the data header row
        c0 = str(row[0] or "").strip()
        c1 = str(row[1] or "").strip()

        if c0 == "Sector" and c1 == "Sub Sector":
            header_found = True
            continue

        if not header_found:
            continue

        # Detect summary timeline rows
        if c1 in ("BAU", "E&P", "High - Ambition", "Target", "Ambitious"):
            vals = [safe_float(v) for v in row[2:6]]
            if len(vals) >= 4:
                timeline[c1] = vals   # [2025, 2030, 2040, 2050]
            continue

        # Parse subsector data rows
        sector    = str(row[0] or "").strip()
        subsector = str(row[1] or "").strip()

        if not sector or sector in ("", "None") or subsector in ("Total Emissions",):
            continue
        if subsector.startswith("Total"):
            continue

        def g(i):
            return safe_float(row[i]) if i < len(row) else 0.0

        subsectors.append({
            "sector":        sector,
            "subsector":     subsector,
            "base":          g(2),
            # 2030
            "bau_2030":      g(3),
            "ep_pct_2030":   g(4),
            "ep_abs_2030":   g(6),
            "amb_pct_2030":  g(7),
            "amb_abs_2030":  g(9),
            # 2040
            "bau_2040":      g(11),
            "ep_pct_2040":   g(12),
            "ep_abs_2040":   g(14),
            "amb_pct_2040":  g(15),
            "amb_abs_2040":  g(17),
            # 2050
            "bau_2050":      g(19),
            "ep_pct_2050":   g(20),
            "ep_abs_2050":   g(22),
            "amb_pct_2050":  g(23),
            "amb_abs_2050":  g(25),
        })

    return subsectors, timeline


def parse_excel(filepath):
    """Master parser — returns complete data dict from ASCENT Excel file."""
    wb = load_workbook(filepath, read_only=True, data_only=True)

    basic        = parse_basic_info(wb)
    dashboard    = parse_dashboard_bau(wb)
    targets      = parse_target_setting(wb)
    er_subs, er_timeline = parse_emission_reduction(wb)

    # Aggregate sector totals from dashboard data
    sector_totals = {}
    for row in dashboard:
        s = row["sector"] or "Other"
        if s not in sector_totals:
            sector_totals[s] = {"base": 0, "y2030": 0, "y2040": 0, "y2050": 0}
        sector_totals[s]["base"]  += row["base"]
        sector_totals[s]["y2030"] += row["y2030"]
        sector_totals[s]["y2040"] += row["y2040"]
        sector_totals[s]["y2050"] += row["y2050"]

    # Grand totals
    def grand(key):
        return sum(v[key] for v in sector_totals.values())

    grand_totals = {
        "base":  grand("base"),
        "y2030": grand("y2030"),
        "y2040": grand("y2040"),
        "y2050": grand("y2050"),
    }

    return {
        "basic":        basic,
        "dashboard":    dashboard,
        "sector_totals":sector_totals,
        "grand_totals": grand_totals,
        "targets":      targets,
        "er_subsectors":er_subs,
        "er_timeline":  er_timeline,
    }


# ═══════════════════════════════════════════════════════
#  ROUTES
# ═══════════════════════════════════════════════════════

@app.route("/", methods=["GET"])
def upload():
    return render_template("upload.html")


@app.route("/upload", methods=["POST"])
def handle_upload():
    if "file" not in request.files:
        flash("No file selected.", "error")
        return redirect(url_for("upload"))

    f = request.files["file"]
    if not f.filename or not f.filename.lower().endswith((".xlsx", ".xlsm")):
        flash("Please upload a valid Excel file (.xlsx or .xlsm).", "error")
        return redirect(url_for("upload"))

    save_path = os.path.join(UPLOAD_FOLDER, "ascent_data.xlsm")
    f.save(save_path)

    try:
        data = parse_excel(save_path)
        session["ascent_data"] = json.dumps(data)
        return redirect(url_for("bau_profile"))
    except Exception as e:
        flash(f"Error parsing Excel file: {str(e)}", "error")
        traceback.print_exc()
        return redirect(url_for("upload"))


def get_data():
    """Load parsed data from session."""
    raw = session.get("ascent_data")
    if not raw:
        return None
    return json.loads(raw)


@app.route("/bau-profile")
def bau_profile():
    data = get_data()
    if not data:
        flash("Please upload an ASCENT file first.", "error")
        return redirect(url_for("upload"))
    return render_template("bau_profile.html", data=data)


@app.route("/target-setting")
def target_setting():
    data = get_data()
    if not data:
        flash("Please upload an ASCENT file first.", "error")
        return redirect(url_for("upload"))
    return render_template("target_setting.html", data=data)


@app.route("/dashboard-bau")
def dashboard_bau():
    data = get_data()
    if not data:
        flash("Please upload an ASCENT file first.", "error")
        return redirect(url_for("upload"))
    return render_template("dashboard_bau.html", data=data)


@app.route("/emission-reduction")
def emission_reduction():
    data = get_data()
    if not data:
        flash("Please upload an ASCENT file first.", "error")
        return redirect(url_for("upload"))
    return render_template("emission_reduction.html", data=data)


@app.route("/api/data")
def api_data():
    data = get_data()
    if not data:
        return jsonify({"error": "No data"}), 404
    return jsonify(data)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
